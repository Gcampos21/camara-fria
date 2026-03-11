from flask import Flask, render_template, request, redirect, jsonify
from openpyxl import load_workbook
from datetime import datetime
import os

app = Flask(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ARQUIVO = os.path.join(BASE_DIR, "estoque.xlsx")


# ----------------------------
# BUSCAR PRODUTO
# ----------------------------

def buscar_produto(codigo):

    wb = load_workbook(ARQUIVO, data_only=True)
    sheet = wb["PRODUTOS"]

    codigo = str(codigo).strip()

    for row in sheet.iter_rows(min_row=2):

        cod_planilha = row[0].value

        if cod_planilha is None:
            continue

        cod_planilha = str(cod_planilha).strip()

        if cod_planilha == codigo:

            produto = row[1].value

            if produto:
                return produto.strip()

    return None


# ----------------------------
# CALCULAR ESTOQUE
# ----------------------------

def calcular_estoque(codigo):

    wb = load_workbook(ARQUIVO)
    sheet = wb["MOVIMENTACOES"]

    estoque = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):

        tipo = row[1]
        cod = row[3]
        qtd = row[5]

        if str(cod).strip() == str(codigo).strip():

            if tipo == "Entrada":
                estoque += qtd

            if tipo == "Saída":
                estoque -= qtd

    return estoque


# ----------------------------
# API BUSCA AUTOMÁTICA
# ----------------------------

@app.route("/produto/<codigo>")
def produto_api(codigo):

    produto = buscar_produto(codigo)

    if not produto:
        return jsonify({"produto": "", "estoque": 0})

    estoque = calcular_estoque(codigo)

    return jsonify({
        "produto": produto,
        "estoque": estoque
    })


# ----------------------------
# MAPA DA CÂMARA
# ----------------------------

@app.route("/mapa")
def mapa():

    wb = load_workbook(ARQUIVO)
    sheet = wb["MOVIMENTACOES"]

    posicoes = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):

        tipo = row[1]
        endereco = row[2]
        produto = row[4]
        qtd = row[5]
        validade = row[6]

        if not endereco:
            continue

        if endereco not in posicoes:

            posicoes[endereco] = {
                "produto": produto,
                "qtd": 0,
                "validade": validade
            }

        if tipo == "Entrada":
            posicoes[endereco]["qtd"] += qtd

        if tipo == "Saída":
            posicoes[endereco]["qtd"] -= qtd

    return render_template("mapa.html", posicoes=posicoes)


# ----------------------------
# CONSULTA
# ----------------------------

@app.route("/consulta")
def consulta():
    return render_template("consulta.html")


@app.route("/buscar", methods=["POST"])
def buscar():

    codigo = request.form["codigo"]

    wb = load_workbook(ARQUIVO)
    sheet = wb["MOVIMENTACOES"]

    estoque_total = 0
    posicoes = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):

        tipo = row[1]
        endereco = row[2]
        cod = row[3]
        qtd = row[5]

        if str(cod).strip() == str(codigo).strip():

            if tipo == "Entrada":

                estoque_total += qtd

                if endereco not in posicoes:
                    posicoes[endereco] = 0

                posicoes[endereco] += qtd

            if tipo == "Saída":

                estoque_total -= qtd

    produto = buscar_produto(codigo)

    return render_template(
        "consulta.html",
        produto=produto,
        codigo=codigo,
        estoque=estoque_total,
        posicoes=posicoes
    )


# ----------------------------
# TELA PRINCIPAL
# ----------------------------

@app.route("/")
def index():
    return render_template("index.html")


# ----------------------------
# MOVIMENTAÇÃO
# ----------------------------

@app.route("/movimentar", methods=["POST"])
def movimentar():

    tipo = request.form["tipo"]
    codigo = request.form["codigo"]

    qtd = request.form.get("quantidade")

    if not qtd:
        return "Informe a quantidade"

    quantidade = int(qtd)

    produto = buscar_produto(codigo)

    if not produto:
        return "Produto não encontrado"

    endereco = ""
    validade = ""

    if tipo == "Entrada":

        setor = request.form["setor"]
        rua = request.form["rua"]
        posicao = request.form["posicao"]
        andar = request.form["andar"]

        endereco = f"{setor}-{rua}-{posicao}-{andar}"

        validade = request.form.get("validade")

        if not validade:
            return "Informe a validade"

    if tipo == "Saída":

        estoque = calcular_estoque(codigo)

        if quantidade > estoque:
            return "Estoque insuficiente"

    wb = load_workbook(ARQUIVO)
    sheet = wb["MOVIMENTACOES"]

    linha = sheet.max_row + 1

    sheet.cell(row=linha, column=1).value = datetime.now()
    sheet.cell(row=linha, column=2).value = tipo
    sheet.cell(row=linha, column=3).value = endereco
    sheet.cell(row=linha, column=4).value = codigo
    sheet.cell(row=linha, column=5).value = produto
    sheet.cell(row=linha, column=6).value = quantidade
    sheet.cell(row=linha, column=7).value = validade

    wb.save(ARQUIVO)

    return redirect("/")


if __name__ == "__main__":
    app.run(debug=True)